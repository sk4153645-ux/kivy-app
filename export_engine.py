# export_engine.py - Complete PDF and Excel Khata Statements
import os
import datetime
from fpdf import FPDF
import openpyxl
import database as db

class ExportEngine:
    @staticmethod
    def generate_farmer_monthly_pdf(farmer_code, start_date, end_date, output_path="statement.pdf"):
        conn = db.get_db()
        farmer = conn.execute("SELECT * FROM farmers WHERE code = ?", (farmer_code,)).fetchone()
        entries = conn.execute(
            "SELECT * FROM milk_purchases WHERE farmer_code = ? AND date BETWEEN ? AND ? ORDER BY date ASC",
            (farmer_code, start_date, end_date)
        ).fetchall()
        payments = conn.execute(
            "SELECT * FROM farmer_payments WHERE farmer_code = ? AND date BETWEEN ? AND ? ORDER BY date ASC",
            (farmer_code, start_date, end_date)
        ).fetchall()
        conn.close()

        if not farmer:
            return False, "Farmer not found."

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", 'B', 16)
        dairy_name = db.get_setting("dairy_name", "Nilgiri Dairy Collection")
        pdf.cell(0, 10, dairy_name, ln=True, align="C")
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 6, f"Farmer Statement: [{farmer['code']}] {farmer['name']} | Phone: {farmer['phone']}", ln=True, align="C")
        pdf.cell(0, 6, f"Period: {start_date} to {end_date}", ln=True, align="C")
        pdf.ln(5)

        # Table Header
        pdf.set_font("Helvetica", 'B', 9)
        pdf.cell(25, 7, "Date", 1)
        pdf.cell(20, 7, "Shift", 1)
        pdf.cell(20, 7, "Type", 1)
        pdf.cell(25, 7, "Litres", 1)
        pdf.cell(20, 7, "Fat", 1)
        pdf.cell(25, 7, "Rate (Rs)", 1)
        pdf.cell(35, 7, "Total (Rs)", 1)
        pdf.ln()

        pdf.set_font("Helvetica", size=9)
        total_litres = 0.0
        total_milk_amt = 0.0
        for e in entries:
            pdf.cell(25, 6, str(e["date"]), 1)
            pdf.cell(20, 6, str(e["shift"]), 1)
            pdf.cell(20, 6, str(e["milk_type"]), 1)
            pdf.cell(25, 6, f"{e['litres']:.2f}", 1)
            pdf.cell(20, 6, f"{e['fat']:.1f}", 1)
            pdf.cell(25, 6, f"{e['rate']:.2f}", 1)
            pdf.cell(35, 6, f"{e['total_amount']:.2f}", 1)
            pdf.ln()
            total_litres += e["litres"]
            total_milk_amt += e["total_amount"]

        pdf.set_font("Helvetica", 'B', 9)
        pdf.cell(65, 7, "Total Milk Supply", 1)
        pdf.cell(25, 7, f"{total_litres:.2f} L", 1)
        pdf.cell(45, 7, "", 1)
        pdf.cell(35, 7, f"Rs.{total_milk_amt:.2f}", 1)
        pdf.ln(10)

        total_paid = sum(p["amount"] for p in payments)
        balance = total_milk_amt - total_paid
        pdf.cell(0, 6, f"Total Milk Amount : Rs. {total_milk_amt:.2f}", ln=True)
        pdf.cell(0, 6, f"Total Payments Settled: Rs. {total_paid:.2f}", ln=True)
        pdf.cell(0, 7, f"Net Balance Payable: Rs. {balance:.2f}", ln=True)

        pdf.output(output_path)
        return True, output_path

    @staticmethod
    def generate_excel_report(output_path="dairy_report.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Today Collection"
        ws.append(["ID", "Date", "Shift", "Farmer Code", "Type", "Litres", "Fat", "SNF", "Rate", "Total Amount"])

        conn = db.get_db()
        rows = conn.execute("SELECT * FROM milk_purchases ORDER BY id DESC").fetchall()
        conn.close()

        for r in rows:
            ws.append([r["id"], r["date"], r["shift"], r["farmer_code"], r["milk_type"], r["litres"], r["fat"], r["snf"], r["rate"], r["total_amount"]])

        wb.save(output_path)
        return True, output_path
